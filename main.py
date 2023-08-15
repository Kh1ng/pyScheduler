import calendar
from datetime import datetime
from openpyxl import Workbook
wb = Workbook()

# Select the worksheet
worksheet = wb.active

#Edit here
year = 2023
month = 9
shifts_per_day = 4
staff = ['Eric', 'Alice', 'Bob', 'Charlie', 'David', 'Joe']

# Write the days to rows 1 (name) & 2 (number)
# Get the days in the month and their abbreviated names
days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
num_days = calendar.monthrange(year, month)
# Write the month to A1
worksheet.cell(row=1, column=1, value=datetime(year, month, 1).strftime('%B'))

# Write the days and their abbreviated names to rows 1 and 2, write month to A1
for i in range(0, num_days[1]):
    worksheet.cell(row=1, column=i + 2,
                   value=(days[i % 7]))
    worksheet.cell(row=2, column=i + 2, value=i + 1)

# Write the names to column A
for i, name in enumerate(staff, start=3):
    worksheet.cell(row=i, column=1, value=name)

# Write the dates to the calendar
totalShifts = num_days[1] * shifts_per_day * len(staff)

# Save the file
wb.save("sample.xlsx")
