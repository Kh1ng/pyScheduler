from datetime import datetime

import openpyxl
import calendar
from openpyxl import Workbook

class Schedule:
    def __init__(self, month, path):
        #attributes
        self.create_schedule(month, path)
        self.month = month
        self.year = 2023
        self.workbook

        #set schedule attributes
        self.total_shifts = calendar.monthrange(2023, 4)
        self.number_doctors = self.nerds.len

        self.filename = month + ".xlsx"
        self.nerds = ['Eric', 'Alice', 'Bob', 'Charlie', 'David', 'Joe']

        # write to excel
        self.write_days_row(self.total_shifts)
        self.write_names_column()


    def __str__(self):
        return f"Schedule: {self.start_time} - {self.end_time} on {self.days}"

    # create a function that takes a month then creates an excel sheet with the schedule for that month
    def create_schedule(self, month, path):
        #create excel sheet
        def create_workbook(path):
            workbook = Workbook()
            workbook.save(path)

        self.workbook = create_workbook("schedule.xlsx")
        self.save()
        self.workbook

    def write_names_column(self):
        # Load the workbook
        workbook = openpyxl.load_workbook(self.filename)

        # Select the worksheet
        worksheet = workbook.active

        # Write the names to column A
        for i, name in enumerate(self.nerds, start=3):
            worksheet.cell(row=i, column=1, value=name)

        # Save the workbook
        workbook.save(self.filename)

    def write_days_row(self, total_shifts):
        # Load the workbook
        workbook = openpyxl.load_workbook(self.filename)

        # Select the worksheet
        worksheet = workbook.active

        # Write the days to rows 1 (name) & 2 (number)
        # Get the days in the month and their abbreviated names
        days = calendar.month_abbr[1:]
        num_days = total_shifts
        days = days[:num_days]

        # Load the workbook
        workbook = openpyxl.load_workbook(self.filename)

        # Select the worksheet
        worksheet = workbook.active

        # Write the days and their abbreviated names to rows 1 and 2, write month to A1
        for i, num_days in enumerate(days, start=2):
            worksheet.cell(row=1, column=i, value=(calendar.day_name[datetime.strptime(('2023-04-' + str(i-1)), '%Y-%m-%d').date().weekday()])[:2])
            worksheet.cell(row=2, column=i, value=i-1)

        self.save()


    # Save the workbook
    def save(self):
        self.workbook.save(self.filename)
        print("Saved to:" + self.path)


